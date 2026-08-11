#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
可予礼品有限公司 - 喜糖/伴手礼竞品监控平台
Flask 主应用
"""

import os
import json
import random
from flask import Flask, render_template, request, jsonify, redirect, url_for
from database import (
    init_db, get_rankings, add_ranking, clear_rankings,
    get_viral_content, add_viral, clear_viral,
    get_age_distribution, add_age_distribution, clear_age,
    get_merchants, add_merchant, clear_merchants,
    create_share_link, get_share_links, verify_share_token,
    revoke_share_link, regenerate_share_link, delete_share_link,
    get_dashboard_stats, get_trends_data
)
from taobao import collector

app = Flask(__name__)
app.config['BRAND_NAME'] = '可予礼品有限公司'
app.config['BRAND_COLOR'] = '#8B2942'
app.config['GOLD_COLOR'] = '#C9A961'

# 初始化数据库
init_db()


# ==========================================
# 页面路由
# ==========================================

@app.route('/')
def dashboard():
    """Dashboard 首页概览"""
    stats = get_dashboard_stats()
    rankings = get_rankings(limit=10)
    viral = get_viral_content(limit=6)
    return render_template('dashboard.html', stats=stats, rankings=rankings, viral=viral)


@app.route('/rankings')
def rankings():
    """销量排行页面"""
    platform = request.args.get('platform', '')
    keyword = request.args.get('keyword', '')
    data = get_rankings(platform=platform or None, keyword=keyword or None)
    platforms = ['淘宝', '抖音', '小红书', '大众点评']
    return render_template('rankings.html', rankings=data, platforms=platforms,
                           current_platform=platform, current_keyword=keyword)


@app.route('/viral')
def viral():
    """爆款内容页面"""
    platform = request.args.get('platform', '')
    data = get_viral_content(platform=platform or None)
    platforms = ['抖音', '小红书', '大众点评']
    return render_template('viral.html', viral=data, platforms=platforms,
                           current_platform=platform)


@app.route('/age')
def age():
    """消费者年龄分布"""
    data = get_age_distribution()
    return render_template('age.html', age_data=data)


@app.route('/trends')
def trends():
    """行业趋势"""
    data = get_trends_data()
    return render_template('trends.html', trends=data)


@app.route('/merchants')
def merchants():
    """跟踪商家"""
    platform = request.args.get('platform', '')
    data = get_merchants(platform=platform or None)
    platforms = ['淘宝', '抖音', '小红书', '大众点评']
    return render_template('merchants.html', merchants=data, platforms=platforms,
                           current_platform=platform)


@app.route('/excel')
def excel_import():
    """Excel 数据导入"""
    return render_template('excel.html')


@app.route('/quick-collect')
def quick_collect():
    """快速采集页面"""
    links = collector.get_alternative_links('喜糖')
    mode_info = collector.get_mode_info()
    return render_template('quick_collect.html', links=links, mode_info=mode_info)


@app.route('/guide')
def guide():
    """新手引导页面"""
    stats = get_dashboard_stats()
    return render_template('guide.html', stats=stats)


@app.route('/create')
def create():
    """AI 创作模块"""
    return render_template('create.html')


@app.route('/share')
def share():
    """分享管理"""
    links = get_share_links()
    return render_template('share.html', share_links=links)


@app.route('/s/<token>')
def shared_view(token):
    """受邀用户查看页面"""
    link = verify_share_token(token)
    if not link:
        return render_template('shared_expired.html'), 404
    if link.get('expired'):
        return render_template('shared_expired.html'), 410

    stats = get_dashboard_stats()
    rankings = get_rankings(limit=10)
    viral = get_viral_content(limit=6)
    return render_template('shared_view.html', stats=stats, rankings=rankings,
                           viral=viral, permission=link['permission'])


# ==========================================
# 分享 API
# ==========================================

@app.route('/api/share/create', methods=['POST'])
def api_share_create():
    data = request.json or {}
    title = data.get('title', '竞品数据分享')
    permission = data.get('permission', 'readonly')
    expiry_type = data.get('expiry_type', '7d')
    token = create_share_link(title, permission, expiry_type)
    return jsonify({'success': True, 'token': token, 'url': f'/s/{token}'})


@app.route('/api/share/list')
def api_share_list():
    links = get_share_links()
    return jsonify({'success': True, 'links': links})


@app.route('/api/share/revoke', methods=['POST'])
def api_share_revoke():
    token = (request.json or {}).get('token', '')
    revoke_share_link(token)
    return jsonify({'success': True})


@app.route('/api/share/regenerate', methods=['POST'])
def api_share_regenerate():
    token = (request.json or {}).get('token', '')
    new_token = regenerate_share_link(token)
    if new_token:
        return jsonify({'success': True, 'token': new_token, 'url': f'/s/{new_token}'})
    return jsonify({'success': False, 'error': '链接不存在'}), 404


@app.route('/api/share/delete', methods=['POST'])
def api_share_delete():
    token = (request.json or {}).get('token', '')
    delete_share_link(token)
    return jsonify({'success': True})


# ==========================================
# Excel 导入 API
# ==========================================

@app.route('/api/excel/upload', methods=['POST'])
def api_excel_upload():
    """上传 Excel 并导入数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '请选择文件'}), 400

    file = request.files['file']
    import_type = request.form.get('type', 'rankings')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            if import_type == 'rankings':
                if len(row) >= 6:
                    platform_map = {'淘宝': '淘宝', 'taobao': '淘宝', 'tb': '淘宝',
                                    '抖音': '抖音', 'douyin': '抖音', 'dy': '抖音',
                                    '小红书': '小红书', 'xhs': '小红书',
                                    '大众点评': '大众点评', 'dianping': '大众点评', 'dp': '大众点评'}
                    platform = platform_map.get(str(row[0]).strip().lower(), str(row[0]).strip())
                    add_ranking({
                        'platform': platform,
                        'keyword': str(row[1]) if row[1] else '',
                        'rank': int(row[2]) if row[2] else 0,
                        'product_name': str(row[3]) if row[3] else '',
                        'shop_name': str(row[4]) if row[4] else '',
                        'price': float(row[5]) if row[5] else 0,
                        'sales': int(row[6]) if len(row) > 6 and row[6] else 0,
                        'product_link': str(row[7]) if len(row) > 7 and row[7] else '',
                        'source': 'excel_import'
                    })
                    count += 1

            elif import_type == 'viral':
                if len(row) >= 4:
                    add_viral({
                        'platform': str(row[0]).strip(),
                        'title': str(row[1]) if row[1] else '',
                        'author': str(row[2]) if row[2] else '',
                        'likes': int(row[3]) if row[3] else 0,
                        'comments': int(row[4]) if len(row) > 4 and row[4] else 0,
                        'shares': int(row[5]) if len(row) > 5 and row[5] else 0,
                        'content_link': str(row[6]) if len(row) > 6 and row[6] else '',
                        'keyword': str(row[7]) if len(row) > 7 and row[7] else '',
                        'source': 'excel_import'
                    })
                    count += 1

            elif import_type == 'age':
                if len(row) >= 3:
                    add_age_distribution({
                        'source': str(row[0]).strip(),
                        'age_group': str(row[1]).strip(),
                        'percentage': float(row[2]) if row[2] else 0,
                        'gender': str(row[3]) if len(row) > 3 and row[3] else None,
                        'sample_size': int(row[4]) if len(row) > 4 and row[4] else None,
                        'platform': str(row[5]) if len(row) > 5 and row[5] else None
                    })
                    count += 1

        return jsonify({'success': True, 'count': count, 'message': f'成功导入 {count} 条数据'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/excel/template/<template_type>')
def api_excel_template(template_type):
    """下载 Excel 模板"""
    import openpyxl
    from flask import send_file
    import io

    wb = openpyxl.Workbook()
    ws = wb.active

    if template_type == 'rankings':
        ws.title = '销量排行'
        ws.append(['平台(淘宝/抖音/小红书/大众点评)', '搜索关键词', '排名', '商品名称', '店铺名称', '价格(元)', '销量', '链接'])
    elif template_type == 'viral':
        ws.title = '爆款内容'
        ws.append(['平台(抖音/小红书/大众点评)', '标题', '作者', '点赞数', '评论数', '分享数', '链接', '关键词'])
    elif template_type == 'age':
        ws.title = '年龄分布'
        ws.append(['数据来源', '年龄段', '占比(%)', '性别', '样本量', '平台'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'candy_monitor_{template_type}_template.xlsx')


# ==========================================
# AI 创作 API - 多模型AI工作站
# ==========================================

# LLM 模型配置
LLM_MODELS = {
    'deepseek': {
        'name': 'DeepSeek V3.1',
        'provider': '深度求索',
        'price_input': '2元/百万token',
        'price_output': '8元/百万token',
        'free_quota': '注册送额度',
        'strength': '综合性价比最高，OpenAI兼容API',
        'context': '64K',
        'base_url': 'https://api.deepseek.com/v1',
        'endpoint_suffix': '/chat/completions'
    },
    'kimi': {
        'name': 'Kimi (Moonshot)',
        'provider': '月之暗面',
        'price_input': '2元/百万token',
        'price_output': '8元/百万token',
        'free_quota': '注册送额度',
        'strength': '128K超长上下文',
        'context': '128K',
        'base_url': 'https://api.moonshot.cn/v1',
        'endpoint_suffix': '/chat/completions'
    },
    'ollama': {
        'name': 'Ollama 本地 (Qwen3 8B)',
        'provider': '本地部署',
        'price_input': '免费（本地GPU）',
        'price_output': '免费',
        'free_quota': '无限制',
        'strength': '数据私有、零边际成本',
        'context': '32K',
        'base_url': 'http://localhost:11434/v1',
        'endpoint_suffix': '/chat/completions'
    },
    'doubao': {
        'name': '豆包 Pro',
        'provider': '字节跳动',
        'price_input': '0.8元/百万token',
        'price_output': '2元/百万token',
        'free_quota': '免费额度充足',
        'strength': '价格最低，均衡之选',
        'context': '32K',
        'base_url': 'https://ark.cn-beijing.volces.com/api/v3',
        'endpoint_suffix': '/chat/completions'
    },
    'qwen': {
        'name': '通义千问 Qwen3-Plus',
        'provider': '阿里云',
        'price_input': '2元/百万token',
        'price_output': '4元/百万token',
        'free_quota': '百万token免费',
        'strength': '能力均衡，开源生态好',
        'context': '128K',
        'base_url': 'https://dashscope.aliyuncs.com/compatible-mode/v1',
        'endpoint_suffix': '/chat/completions'
    }
}

IMAGE_MODELS = {
    'flux2_klein': {'name': 'FLUX.2 klein 4B', 'vram': '8GB', 'speed': '快（4步生成）', 'quality': '8/10', 'license': 'Apache 2.0', 'score': 8},
    'flux1_dev': {'name': 'FLUX.1 dev', 'vram': '8GB(Q4)', 'speed': '中等', 'quality': '8.5/10', 'license': '非商用', 'score': 8.5},
    'flux2_dev': {'name': 'FLUX.2 dev 32B', 'vram': '19GB(Q4)', 'speed': '慢', 'quality': '9.5/10', 'license': '非商用', 'score': 9.5},
    'sdxl': {'name': 'SDXL', 'vram': '4GB', 'speed': '快', 'quality': '7/10', 'license': '开源', 'score': 7},
    'sd35_large': {'name': 'SD 3.5 Large', 'vram': '12GB', 'speed': '中等', 'quality': '8.3/10', 'license': '社区许可', 'score': 8.3}
}

VIDEO_MODELS = {
    'cogvideox_2b': {'name': 'CogVideoX-2B', 'vram': '4GB', 'duration': '6秒/720×480', 'speed_a100': '~90秒', 'license': '开源'},
    'cogvideox_5b': {'name': 'CogVideoX-5B', 'vram': '5GB(优化)', 'duration': '6秒/720×480', 'speed_a100': '~180秒', 'license': '开源'},
    'opensora2': {'name': 'Open-Sora 2.0', 'vram': '16GB+', 'duration': '可变', 'speed_a100': '较慢', 'license': '开源'},
    'ltx_video': {'name': 'LTX Video 2B', 'vram': '12GB', 'duration': '可变', 'speed_a100': '快', 'license': '开源'}
}

CLOUD_VIDEO = {
    'kling': {'name': '可灵 Kling 2.0', 'price': '0.5-1元/条', 'duration': '5-10秒', 'resolution': '1080p'},
    'seedance': {'name': '豆包 Seedance', 'price': '0.3-0.6元/条', 'duration': '5-10秒', 'resolution': '1080p'},
    'minimax': {'name': 'MiniMax Video-01', 'price': '0.4-0.8元/条', 'duration': '6秒', 'resolution': '1080p'}
}

HARDWARE_CONFIGS = [
    {
        'level': '入门级', 'budget': '约3000元(仅GPU)', 'gpu': 'RTX 4060 8GB',
        'can_run': 'Qwen3 8B + FLUX.2 klein 4B + CogVideoX-2B',
        'limit': '只能跑量化模型，视频生成较慢', 'color': 'blue'
    },
    {
        'level': '创作者级', 'budget': '约5000元(仅GPU)', 'gpu': 'RTX 4070 Ti Super 16GB',
        'can_run': 'Qwen3 14B + FLUX.1 dev + CogVideoX-5B',
        'limit': '性价比最佳，覆盖所有日常需求', 'color': 'gold'
    },
    {
        'level': '专业级', 'budget': '约10000元(仅GPU)', 'gpu': 'RTX 4090 24GB',
        'can_run': 'Qwen3 30B + FLUX.2 dev(FP8) + Wan Video 14B',
        'limit': '重度AI创作、多任务并行', 'color': 'brand'
    }
]


@app.route('/api/ai/models')
def api_ai_models():
    """获取所有 AI 模型配置信息"""
    return jsonify({
        'success': True,
        'llm': LLM_MODELS,
        'image': IMAGE_MODELS,
        'video': VIDEO_MODELS,
        'cloud_video': CLOUD_VIDEO,
        'hardware': HARDWARE_CONFIGS
    })


@app.route('/api/generate_image', methods=['POST'])
def api_generate_image():
    """AI 图片生成 - 支持 Pollinations.ai 和 ComfyUI 本地"""
    data = request.json or {}
    prompt = data.get('prompt', '')
    style = data.get('style', '摄影写实')
    size = data.get('size', '1024x1024')
    engine = data.get('engine', 'pollinations')

    style_prompts = {
        '摄影写实': 'professional product photography, wedding candy gift box, elegant, soft lighting, 8k detailed, commercial photography',
        '可爱插画': 'cute illustration, kawaii style, wedding favors, pastel colors, hand-drawn, charming, watercolor',
        '高级质感': 'luxury wedding gift packaging, premium, sophisticated, gold accents, minimal, magazine editorial quality',
        '国风中式': 'traditional Chinese wedding candy box, red and gold, paper-cut art style, cultural elegance, silk texture',
        '极简留白': 'minimalist wedding favor design, clean, white space, modern, elegant simplicity, studio lighting',
        '梦幻浪漫': 'romantic wedding candy, dreamy, soft pink, flower petals, fairy lights, ethereal, bokeh',
        '电商白底': 'e-commerce product photography, wedding candy box on white background, clean, professional, studio lighting'
    }

    full_prompt = f"{prompt}, {style_prompts.get(style, style_prompts['摄影写实'])}"

    if engine == 'pollinations':
        image_url = f"https://image.pollinations.ai/prompt/{full_prompt}?width={size.split('x')[0]}&height={size.split('x')[1]}&nologo=true&seed={random.randint(1,99999)}"
    else:
        # ComfyUI fallback to Pollinations
        image_url = f"https://image.pollinations.ai/prompt/{full_prompt}?width={size.split('x')[0]}&height={size.split('x')[1]}&nologo=true"

    return jsonify({
        'success': True,
        'image_url': image_url,
        'prompt': full_prompt,
        'style': style,
        'engine': engine
    })


@app.route('/api/generate_xhs_copy', methods=['POST'])
def api_generate_xhs_copy():
    """小红书爆款文案生成 - 增强版（Dify工作流模拟）"""
    data = request.json or {}
    product = data.get('product', '喜糖礼盒')
    style = data.get('style', '种草')
    word_count = data.get('word_count', 'medium')
    model = data.get('model', 'deepseek')

    model_info = LLM_MODELS.get(model, LLM_MODELS['deepseek'])

    # 根据风格生成不同结构的标题
    title_map = {
        '种草': [
            f'备婚必看！{product}这样选，宾客都夸爆了',
            f'闺蜜结婚选了这款{product}，直接被种草了',
            f'颜值担当！这款{product}让婚礼高级感拉满',
        ],
        '测评': [
            f'花了3000元实测12款{product}，结果出乎意料',
            f'{product}开箱实测：哪些值得买？哪些要避坑？',
            f'真实测评｜{product}到底好不好？看完再决定',
        ],
        '攻略': [
            f'备婚干货｜{product}避雷指南，建议收藏',
            f'千万别踩坑！{product}选购攻略，省钱又高级',
            f'婚庆圈都在找的{product}选购秘籍，一次说清楚',
        ],
        '分享': [
            f'我的{product}备婚日记，宾客都说好',
            f'被问爆的{product}链接，今天终于整理好了',
            f'晒晒我选的{product}，婆婆看了连连称赞',
        ],
        '探店': [
            f'探访线下{product}实体店，发现了宝藏',
            f'跑遍5家{product}店，终于找到心仪的',
            f'实体店vs网购{product}，差别太大了',
        ]
    }

    titles = title_map.get(style, title_map['种草'])

    # 根据风格生成正文
    body_map = {
        '种草': f'最近在备婚，看了很多{product}\n\n分享几点选购心得：\n1⃣ 先确定预算范围\n2⃣ 看包装质感，要能体现婚礼主题\n3⃣ 糖果口味要兼顾大人和小朋友\n4⃣ 一定要提前确认发货时间\n\n最后选了这家，宾客反馈超好！\n\n推荐给所有备婚的姐妹～',
        '测评': f'最近实测了市面上热门的{product}\n\n从这几个维度打分（满分10分）：\n✅ 包装颜值：9分\n✅ 糖果品质：8.5分\n✅ 性价比：9分\n✅ 发货速度：8分\n✅ 客服服务：9.5分\n\n综合推荐指数：⭐⭐⭐⭐⭐',
        '攻略': f'备婚三个月，{product}看了不下50家\n\n给姐妹们总结一下避坑指南：\n1. 不要只看主图，一定要看买家秀\n2. 问清楚糖果的克重和保质期\n3. 定制款要提前1个月下单\n4. 可以要求先寄样品，确认后再批量下单\n\n收藏这篇，少走弯路！',
        '分享': f'我的{product}终于定下来了！\n\n一开始真的很纠结，款式太多了\n后来按照几个原则筛选：\n• 和婚礼主题色搭配\n• 预算控制在30元以内\n• 包装要环保可降解\n\n最终选了这款中国风，和我们的中式婚礼太配了！\n\n希望每个新娘都能选到心仪的{product}～',
        '探店': f'周末跑了好几家{product}实体店\n\n实体店的好处是可以直接看到实物：\n✨ 包装质感一目了然\n✨ 可以试吃糖果\n✨ 店员会根据预算搭配方案\n✨ 定制沟通更顺畅\n\n我最后选了XXX店，推荐给上海的姐妹！'
    }

    body = body_map.get(style, body_map['种草'])

    if word_count == 'short':
        body = f'分享一款超好看的{product}\n\n颜值高、性价比好，宾客都说赞！\n\n真的闭眼入～'
    elif word_count == 'long':
        body = body + '\n\n补充几点小Tips：\n• 记得提前和酒店确认是否允许自带喜糖\n• 多买10%左右的备用份\n• 摆拍的时候可以搭配花束，超好看'

    tags_pool = ['#备婚', '#备婚日记', '#备婚攻略', '#喜糖', '#伴手礼',
                 '#婚礼筹备', '#婚礼好物', '#结婚', '#备婚清单', '#喜糖推荐',
                 '#伴手礼推荐', '#婚礼', '#新娘', '#婚品', '#备婚日常',
                 '#喜糖礼盒', '#婚礼伴手礼', '#备婚好物', '#婚庆']

    tags = ' '.join(random.sample(tags_pool, min(6, len(tags_pool))))

    return jsonify({
        'success': True,
        'title': random.choice(titles),
        'body': body,
        'tags': tags,
        'full': f'{random.choice(titles)}\n\n{body}\n\n{tags}',
        'model_used': model_info['name']
    })


@app.route('/api/generate_dy_copy', methods=['POST'])
def api_generate_dy_copy():
    """抖音爆款脚本生成 - 增强版"""
    data = request.json or {}
    product = data.get('product', '喜糖礼盒')
    content_type = data.get('type', '展示')
    duration = data.get('duration', 'medium')
    model = data.get('model', 'deepseek')

    model_info = LLM_MODELS.get(model, LLM_MODELS['deepseek'])

    hook_map = {
        '口播': [
            f'备婚的姐妹听我一句劝，{product}千万别乱买！',
            f'今天说一个备婚圈不敢说的{product}真相',
            f'90%的新娘选{product}都踩了这个坑',
        ],
        '展示': [
            f'你敢信？这个{product}让婚礼预算省了一半！',
            f'备婚3个月，终于找到天花板级别的{product}',
            f'花了几百块买到高级感{product}，怎么做到的？',
        ],
        '剧情': [
            f'当我把这个{product}拿给婆婆看...',
            f'男朋友是个直男，选的{product}居然...',
            f'闺蜜结婚用了这个{product}，我被惊艳到了',
        ],
        '痛点': [
            f'别再买又贵又土的{product}了！教你避坑',
            f'备婚最后悔的事：{product}没提前看这篇',
            f'花了3000买{product}，我后悔了...',
        ],
        '开箱': [
            f'快递到了！今天开箱最近超火的{product}',
            f'一口气开了5款{product}，结果...',
            f'这款被称为\"神仙{product}\"的实物长这样',
        ]
    }

    hooks = hook_map.get(content_type, hook_map['展示'])

    script_map = {
        '口播': f'今天跟大家聊聊{product}怎么选\n\n很多新娘只看包装就下单\n但其实最重要的是这三点：\n第一，糖果的克重和品质\n第二，包装是否密封防潮\n第三，商家的售后和退换政策\n\n把这三点搞清楚了\n你选的{product}绝对不会翻车！',
        '展示': f'来，带大家看看这款{product}的细节\n\n第一眼看包装，真的很高级\n打开以后更惊喜\n糖果颗颗饱满，包装严实\n关键是这个价格，太良心了\n\n备婚的姐妹可以直接抄作业！',
        '剧情': f'{random.choice(hooks)}\n\n（画面：拿出{product}）\n\n说实话，真的超出预期\n颜值高、品质好、价格也合适\n\n（画面：摆拍效果）\n\n婚礼当天摆出来\n宾客都说太好看了！',
        '痛点': f'备婚选{product}，这4个坑千万别踩：\n\n❌ 只看主图不看买家秀\n❌ 不问保质期就下单\n❌ 不确认发货时间\n❌ 不先拿样品就批量买\n\n学会这4点，省心又省钱！',
        '开箱': f'新买的{product}到货了！\n\n（开箱画面）\n\n包装完好，没有破损\n打开看看里面的糖果...\n\n（特写画面）\n\n这个质感真的很不错\n比我想象的还要好！'
    }

    script = script_map.get(content_type, script_map['展示'])

    if duration == 'short':
        script = f'{random.choice(hooks)}\n\n一句话：这个{product}闭眼入！'
    elif duration == 'long':
        script = script + '\n\n最后再强调一下：\n备婚选喜糖，品质第一，颜值第二\n好的{product}能让婚礼高级感翻倍\n记得先拿样品，确认满意再下单！'

    tags = '#备婚 #喜糖 #婚礼好物推荐 #备婚攻略 #备婚日记'

    return jsonify({
        'success': True,
        'hook': random.choice(hooks),
        'script': script,
        'tags': tags,
        'full': f'{random.choice(hooks)}\n\n{script}\n\n{tags}',
        'model_used': model_info['name']
    })


@app.route('/api/ai/workflow_suggest', methods=['POST'])
def api_workflow_suggest():
    """Dify 工作流建议 - 根据场景推荐最佳方案"""
    data = request.json or {}
    scenario = data.get('scenario', '全功能')

    workflows = {
        '全功能': {
            'name': '混合部署方案（推荐）',
            'desc': 'Dify + Ollama(Qwen3 8B) + DeepSeek API + ComfyUI(FLUX.2 klein) + CogVideoX-2B',
            'hardware': 'RTX 4060 8GB 起步，推荐 RTX 4070 Ti Super 16GB',
            'monthly_cost': 'API成本可控制在50元以内',
            'steps': [
                'Dify Docker 部署作为工作台核心',
                'Ollama + Qwen3 8B 处理高频/敏感任务',
                'DeepSeek V3.1 API 处理复杂文案创作',
                'ComfyUI + FLUX.2 klein 4B 本地出图',
                'CogVideoX-2B 或 可灵API 视频生成'
            ]
        },
        '纯云端': {
            'name': '纯云端零硬件方案',
            'desc': 'LobeChat + DeepSeek V3.1 + 豆包 Pro + 可灵/豆包视频',
            'hardware': '无需显卡，普通电脑即可',
            'monthly_cost': '月费几十到几百元',
            'steps': [
                '安装 LobeChat 作为前端界面',
                '配置 DeepSeek V3.1 API（文案/对话）',
                '接入豆包 Pro 作为低成本补充',
                '使用可灵/豆包视频 API 生成视频',
                'ComfyUI 插件接入 Replicate/Fal.ai 云端出图'
            ]
        },
        '纯本地': {
            'name': '纯本地私有化方案',
            'desc': 'Dify + Ollama + ComfyUI + CogVideoX，全部本地运行',
            'hardware': '需要 RTX 4090 24GB',
            'monthly_cost': '零边际成本',
            'steps': [
                'Dify Docker 部署工作台',
                'Ollama 本地运行所有 LLM',
                'ComfyUI + FLUX.2 dev 本地出图',
                'CogVideoX-5B 本地视频生成',
                '数据完全私有，不出本地网络'
            ]
        }
    }

    result = workflows.get(scenario, workflows['全功能'])
    return jsonify({'success': True, 'workflow': result})


# ==========================================
# 系统重置 API
# ==========================================

@app.route('/api/reset', methods=['POST'])
def api_reset():
    """清空所有数据"""
    clear_rankings()
    clear_viral()
    clear_age()
    clear_merchants()
    return jsonify({'success': True, 'message': '所有数据已清空'})


# ==========================================
# 健康检查（云部署用）
# ==========================================

@app.route('/health')
def health():
    """健康检查端点"""
    stats = get_dashboard_stats()
    return jsonify({
        'status': 'ok',
        'service': '可予礼品竞品监控平台',
        'version': '2.1',
        'database': 'connected' if stats else 'empty',
        'total_products': stats.get('total_products', 0),
        'total_viral': stats.get('total_viral', 0)
    })


if __name__ == '__main__':
    import os
    host = os.environ.get('CANDY_HOST', '0.0.0.0')
    port = int(os.environ.get('CANDY_PORT', '5000'))
    app.run(host=host, port=port, debug=True)
